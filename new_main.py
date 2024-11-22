from PyQt6 import uic
from PyQt6.QtCore import pyqtSignal
from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QTableWidgetItem, QWidget, QVBoxLayout, \
    QPushButton
import openpyxl as xl

Form, _ = uic.loadUiType("untitled.ui")
FirstWindowForm, _ = uic.loadUiType("untitled1.ui")
SecondWindowForm, _ = uic.loadUiType("untitled2.ui")


class Ui(QMainWindow, Form):
    def __init__(self):
        super(Ui, self).__init__()
        self.setupUi(self)

        # Подключение кнопок
        self.pushButton.clicked.connect(self.btpress)
        self.pushButton_2.clicked.connect(self.open_first_ui_window)
        self.pushButton_3.clicked.connect(self.open_second_ui_window)

        # Обработка изменений в таблице
        self.tableWidget.itemChanged.connect(self.on_item_changed)

        # Глобальная переменная для данных
        self.data_table = []

        # Инициализация дополнительных окон
        self.first_window = FirstUIWindow(self.data_table)
        self.second_window = SecondUIWindow(self.data_table)
        self.first_window.closed.connect(self.unlock_table)
        self.second_window.closed.connect(self.unlock_table)

    def btpress(self):
        data_table = self.open_file_dialog()
        print(data_table)
        if data_table:  # Если данные успешно загружены
            self.data_table = data_table
            print(f"Обновлённая data_table: {self.data_table}")
        else:
            # Очистка таблицы
            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(0)

    def open_first_ui_window(self):
        print("Открытие первого окна")
        self.lock_table()
        self.first_window.modify_data_table(self.data_table)
        self.first_window.show()

    def open_second_ui_window(self):
        print("Открытие второго окна")
        self.second_window.modify_data_table(self.data_table)
        self.second_window.show()
        self.lock_table()

    def lock_table(self):
        """Блокирует таблицу для редактирования."""
        self.tableWidget.setEnabled(False)
        self.pushButton.setEnabled(False)

    def unlock_table(self):
        """Разблокирует таблицу после закрытия окон."""
        self.tableWidget.setEnabled(True)
        self.pushButton.setEnabled(True)

    def tabl(self, way: str) -> list:
        # Загрузка данных из Excel
        table = xl.open(way).active
        row = []
        self.tableWidget.blockSignals(True)
        self.table_filling_in_progress = True
        self.tableWidget.setColumnCount(table.max_column)
        self.tableWidget.setRowCount(table.max_row)
        for i in range(table.max_column):
            a = []
            for j in range(1, table.max_row + 1):
                if table[j][i].value:
                    self.update_cell(j - 1, i, table[j][i].value)
                    a.append(table[j][i].value)
            row.append(a)
        self.table_filling_in_progress = False
        self.tableWidget.blockSignals(False)
        return row

    def update_cell(self, row: int, column: int, value):
        # Обновление ячейки в таблице
        item = QTableWidgetItem(str(value))
        self.tableWidget.setItem(row, column, item)

    def open_file_dialog(self) -> list:
        # Диалоговое окно для выбора файла
        file_path, _ = QFileDialog.getOpenFileName(self, "Открыть файл", "", "Файл Excel (*.xlsx;*.xlx)")
        if file_path:
            return self.tabl(file_path)
        else:
            QMessageBox.warning(self, "Предупреждение", "Не удалось открыть файл")
            return []

    def on_item_changed(self, item):
        # Обработка изменений ячейки таблицы
        if self.table_filling_in_progress:
            return
        new_value = item.text()
        row = item.column()
        column = item.row()
        if not self.is_number(new_value):
            QMessageBox.warning(self, "Ошибка ввода", "Вводите только числовые значения!")
            self.tableWidget.blockSignals(True)
            previous_value = self.data_table[row][column] if column < len(self.data_table[row]) else ""
            item.setText(str(previous_value))
            self.tableWidget.blockSignals(False)
            return
        new_value = int(new_value)
        if row < len(self.data_table) and column < len(self.data_table[row]):
            self.data_table[row][column] = new_value
        else:
            while len(self.data_table) <= row:
                self.data_table.append([])
            while len(self.data_table[row]) <= column:
                self.data_table[row].append("")
            self.data_table[row][column] = new_value

    def is_number(self, value: str) -> bool:
        try:
            float(value)
            return True
        except ValueError:
            return False


class FirstUIWindow(QWidget, FirstWindowForm):
    closed = pyqtSignal()

    def __init__(self, data_table):
        super(FirstUIWindow, self).__init__()
        self.setupUi(self)
        self.fl = True
        self.data_table = data_table

        self.container_widget = self.findChild(QWidget, "buttonContainer")
        self.button_layout = self.container_widget.layout()

        self.container_widget_2 = self.findChild(QWidget, "buttonContainer_2")
        self.button_layout_2 = self.container_widget_2.layout()

    def rang(self, data_table):
        self.clear_buttons()
        print('///////////')
        print(self.data_table)
        if len(data_table) == 2:
            self.add_button(self.button_layout, 'Стьюдент зав', 300, 30)
            self.add_button(self.button_layout, 'Стьюдент незав', 300, 30)
            self.add_button(self.button_layout, 'Манна-Уитни', 300, 30)
            self.add_button(self.button_layout, 'Стьюдент зав', 300, 30)
            self.add_button(self.button_layout, 'Вилкоксона', 300, 30)
        if len(data_table) > 2:
            self.add_button(self.button_layout, 'Круаска Уолиса', 300, 30)
            self.add_button(self.button_layout_2, 'Стьюдент зав', 300, 30)
            self.add_button(self.button_layout_2, 'Стьюдент незав', 300, 30)
            self.add_button(self.button_layout_2, 'Манна-Уитни', 300, 30)
            self.add_button(self.button_layout_2, 'Стьюдент зав', 300, 30)
            self.add_button(self.button_layout_2, 'Вилкоксона', 300, 30)

    def add_button(self, button_layout, label, width=None, height=None):
        button = QPushButton(label)

        # Установка размеров кнопки, если указаны
        if width and height:
            button.setFixedSize(width, height)

        # Добавляем кнопку в макет
        button_layout.addWidget(button)

    def closeEvent(self, event):
        """Отправляет сигнал при закрытии окна."""
        self.closed.emit()
        super().closeEvent(event)

    def clear_buttons(self):
        # Перебираем все элементы в макете
        while self.button_layout.count():
            item = self.button_layout.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()

    def modify_data_table(self, data_table):
        text = f'Число дисперсий в ваших данных = {len(data_table)} \n\n  Вам подходят следующие критерии:'
        print(text)
        self.textEdit.setPlainText(text)  # Если это QTextEdit
        min_width = len(f'Число дисперсий в ваших данных = {len(data_table)}') * 10
        self.textEdit.setMinimumWidth(min_width)
        self.rang(data_table)


class SecondUIWindow(QWidget, SecondWindowForm):
    closed = pyqtSignal()

    def __init__(self, data_table):
        super(SecondUIWindow, self).__init__()
        self.setupUi(self)
        self.fl = True
        self.data_table = data_table

        self.container_widget = self.findChild(QWidget, "buttonContainer")
        self.button_layout = self.container_widget.layout()

        self.container_widget_2 = self.findChild(QWidget, "buttonContainer_2")
        self.button_layout_2 = self.container_widget_2.layout()

    def rang(self, data_table):
        self.clear_buttons()
        print('///////////')
        print(self.data_table)
        if len(data_table) == 2:
            self.add_button(self.button_layout, 'Фишера', 300, 30)
        if len(data_table) > 2:
            self.add_button(self.button_layout, 'Бартлета', 300, 30)
            self.add_button(self.button_layout, 'Левене', 300, 30)
            self.add_button(self.button_layout_2, 'Фишера', 300, 30)


    def add_button(self, button_layout, label, width=None, height=None):
        button = QPushButton(label)

        # Установка размеров кнопки, если указаны
        if width and height:
            button.setFixedSize(width, height)

        # Добавляем кнопку в макет
        button_layout.addWidget(button)

    def closeEvent(self, event):
        """Отправляет сигнал при закрытии окна."""
        self.closed.emit()
        super().closeEvent(event)

    def clear_buttons(self):
        # Перебираем все элементы в макете
        while self.button_layout.count():
            item = self.button_layout.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()

    def modify_data_table(self, data_table):
        text = f'Число дисперсий в ваших данных = {len(data_table)} \n\n  Вам подходят следующие критерии:'
        print(text)
        self.textEdit.setPlainText(text)  # Если это QTextEdit
        min_width = len(f'Число дисперсий в ваших данных = {len(data_table)}') * 10
        self.textEdit.setMinimumWidth(min_width)
        self.rang(data_table)


if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    w = Ui()
    w.show()
    sys.exit(app.exec())
