from PyQt6 import uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QTableWidgetItem, QWidget
import openpyxl as xl

# Загружаем UI файл и создаем классы для формы и окна
Form, _ = uic.loadUiType("untitled.ui")
FirstWindowForm, _ = uic.loadUiType("untitled1.ui")
SecondWindowForm, _ = uic.loadUiType("untitled2.ui")


class Ui(QMainWindow, Form):
    def __init__(self):
        super(Ui, self).__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.btpress)
        self.pushButton_2.clicked.connect(self.open_first_ui_window)
        self.pushButton_3.clicked.connect(self.open_second_ui_window)

        # Подписываемся на сигнал itemChanged, который отслеживает изменения в ячейках таблицы
        # Подписываемся на сигнал itemChanged, который отслеживает изменения в ячейках таблицы
        self.tableWidget.itemChanged.connect(self.on_item_changed)

        self.data_table = []  # Список для хранения данных таблицы
        self.table_filling_in_progress = False  # Флаг для отслеживания состояния заполнения таблицы

        self.first_window = FirstUIWindow(self.data_table)
        self.second_window = SecondUIWindow(self.data_table)

    def btpress(self):
        data_table = self.open_file_dialog()
        print(data_table)
        if data_table:  # Проверка на наличие данных
            self.data_table = data_table
        else:
            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(0)

    def open_first_ui_window(self):
        self.first_window.show()

    def open_second_ui_window(self):
        self.second_window.show()

    def tabl(self, way: str) -> list:
        table = xl.open(way).active
        row = []
        self.tableWidget.blockSignals(True)
        self.table_filling_in_progress = True  # Устанавливаем флаг заполнения таблицы
        self.tableWidget.setColumnCount(table.max_column)
        self.tableWidget.setRowCount(table.max_row)
        for i in range(table.max_column):
            a = []
            for j in range(1, table.max_row + 1):
                if table[j][i].value:
                    print(i, j, table[j][i].value)
                    self.update_cell(j - 1, i, table[j][i].value)  # Важно: row и column в правильном порядке
                    a.append(table[j][i].value)
            row.append(a)
        self.table_filling_in_progress = False  # Сбрасываем флаг после завершения
        self.tableWidget.blockSignals(False)  # Включаем сигналы обратно
        return row

    def update_cell(self, row: int, column: int, value):
        item = QTableWidgetItem(str(value))
        self.tableWidget.setItem(row, column, item)

    def open_file_dialog(self) -> list:
        # Убираем создание нового приложения и main_window
        file_path, _ = QFileDialog.getOpenFileName(self, "Открыть файл", "", "Файл Excel (*.xlsx;*.xlx)")

        if file_path:
            return self.tabl(file_path)
        else:
            QMessageBox.warning(self, "Предупреждение", "Не удалось открыть файл")
            return []

    # Обработчик изменения ячейки таблицы
    def on_item_changed(self, item):
        if self.table_filling_in_progress:
            return
        new_value = item.text()
        row = item.column()
        column = item.row()
        if not self.is_number(new_value):
            QMessageBox.warning(self, "Ошибка ввода", "Вводите только числовые значения!")
            # Не очищаем ячейку, просто восстанавливаем предыдущее значение
            self.tableWidget.blockSignals(True)  # Блокируем сигнал, чтобы избежать рекурсии
            # Устанавливаем предыдущее значение
            previous_value = self.data_table[row][column] if column < len(self.data_table[row]) else ""
            item.setText(str(previous_value))  # Восстанавливаем предыдущее значение
            self.tableWidget.blockSignals(False)  # Включаем сигнал обратно
            return

        print(row, column)
        print(self.data_table)
        new_value = int(item.text())  # Получаем новое значение

        # Обновляем список данных
        if row < len(self.data_table) and column < len(self.data_table[row]):
            self.data_table[row][column] = new_value
            print(f"Значение в ячейке ({row}, {column}) изменено на: {new_value}")
            print(self.data_table)

        # Если таблица изменена, но соответствующего значения нет в списке, можно добавить элемент
        else:
            while len(self.data_table) <= row:
                self.data_table.append([])

            while len(self.data_table[row]) <= column:
                self.data_table[row].append("")

            self.data_table[column][row] = new_value

            print(f"Добавлено новое значение в список данных: {new_value}")
            print(self.data_table)

    def is_number(self, value: str) -> bool:
        try:
            float(value)  # Пробуем преобразовать строку в число
            return True
        except ValueError:
            return False

    def generalPOP(self):
        return

    def dispersion(self):
        return


class FirstUIWindow(QWidget, FirstWindowForm):
    def __init__(self, data_table):

        super(FirstUIWindow, self).__init__()
        self.setupUi(self)
        self.data_table = data_table
        print(data_table)


class SecondUIWindow(QWidget, SecondWindowForm):
    def __init__(self, data_table):
        super(SecondUIWindow, self).__init__()
        self.setupUi(self)
        self.data_table = data_table
        print(data_table)


if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    w = Ui()
    w.show()
    sys.exit(app.exec())
