import openpyxl as xl
from PyQt6 import uic
from PyQt6.QtCore import pyqtSignal
from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QTableWidgetItem, QWidget, QPushButton
import sys
import os

# Определим базовый путь в зависимости от того, запущен ли файл как .exe или нет
if getattr(sys, 'frozen', False):
    # Если приложение собрано (frozen), используем специальный путь
    base_path = sys._MEIPASS
else:
    # Если приложение запущено в режиме разработки, используем текущую директорию
    base_path = os.path.abspath(".")
Form, _ = uic.loadUiType(os.path.join(base_path, "web_ui", "untitled.ui"))
FirstWindowForm, _ = uic.loadUiType(os.path.join(base_path, "web_ui", "untitled1.ui"))
SecondWindowForm, _ = uic.loadUiType(os.path.join(base_path, "web_ui", "untitled2.ui"))


class Ui(QMainWindow, Form):
    def __init__(self, handler):
        super(Ui, self).__init__()
        self.setupUi(self)
        self.handler = handler
        # Подключение кнопок
        self.pushButton.clicked.connect(self.btpress)
        self.pushButton_2.clicked.connect(self.open_first_ui_window)
        self.pushButton_3.clicked.connect(self.open_second_ui_window)

        # Обработка изменений в таблице
        self.tableWidget.itemChanged.connect(self.on_item_changed)

        # Глобальная переменная для данных
        self.data_table = []

        # Инициализация дополнительных окон
        self.first_window = FirstUIWindow(self.data_table, self.handler)
        self.second_window = SecondUIWindow(self.data_table)
        self.first_window.closed.connect(self.unlock_table)
        self.second_window.closed.connect(self.unlock_table)

    def btpress(self):
        data_table = self.open_file_dialog()
        print(data_table)
        if data_table:  # Если данные успешно загружены
            self.data_table = data_table
            print(f"Обновлённая data_table: {self.data_table}")
        else:
            # Очистка таблицы
            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(0)

    def open_first_ui_window(self):
        print("Открытие первого окна")
        self.lock_table()
        self.first_window.modify_data_table(self.data_table)
        self.first_window.show()

    def open_second_ui_window(self):
        print("Открытие второго окна")
        self.second_window.modify_data_table(self.data_table)
        self.second_window.show()
        self.lock_table()

    def lock_table(self):
        """Блокирует таблицу для редактирования."""
        self.tableWidget.setEnabled(False)
        self.pushButton.setEnabled(False)

    def unlock_table(self):
        """Разблокирует таблицу после закрытия окон."""
        self.tableWidget.setEnabled(True)
        self.pushButton.setEnabled(True)

    def tabl(self, way: str) -> list:
        # Загрузка данных из Excel
        table = xl.open(way).active
        row = []
        self.tableWidget.blockSignals(True)
        self.table_filling_in_progress = True
        self.tableWidget.setColumnCount(table.max_column)
        self.tableWidget.setRowCount(table.max_row)
        for i in range(table.max_column):
            a = []
            for j in range(1, table.max_row + 1):
                if table[j][i].value:
                    self.update_cell(j - 1, i, table[j][i].value)
                    a.append(table[j][i].value)
            row.append(a)
        self.table_filling_in_progress = False
        self.tableWidget.blockSignals(False)
        return row

    def update_cell(self, row: int, column: int, value):
        # Обновление ячейки в таблице
        item = QTableWidgetItem(str(value))
        self.tableWidget.setItem(row, column, item)

    def open_file_dialog(self) -> list:
        # Диалоговое окно для выбора файла
        file_path, _ = QFileDialog.getOpenFileName(self, "Открыть файл", "", "Файл Excel (*.xlsx;*.xlx)")
        if file_path:
            return self.tabl(file_path)
        else:
            QMessageBox.warning(self, "Предупреждение", "Не удалось открыть файл")
            return []

    def on_item_changed(self, item):
        # Обработка изменений ячейки таблицы
        if self.table_filling_in_progress:
            return
        new_value = item.text()
        row = item.column()
        column = item.row()
        if not self.is_number(new_value):
            QMessageBox.warning(self, "Ошибка ввода", "Вводите только числовые значения!")
            self.tableWidget.blockSignals(True)
            previous_value = self.data_table[row][column] if column < len(self.data_table[row]) else ""
            item.setText(str(previous_value))
            self.tableWidget.blockSignals(False)
            return
        new_value = int(new_value)
        if row < len(self.data_table) and column < len(self.data_table[row]):
            self.data_table[row][column] = new_value
        else:
            while len(self.data_table) <= row:
                self.data_table.append([])
            while len(self.data_table[row]) <= column:
                self.data_table[row].append("")
            self.data_table[row][column] = new_value

    def is_number(self, value: str) -> bool:
        try:
            float(value)
            return True
        except ValueError:
            return False


class FirstUIWindow(QWidget, FirstWindowForm):
    closed = pyqtSignal()

    def __init__(self, data_table, handler):
        super(FirstUIWindow, self).__init__()
        self.hanler = handler
        self.setupUi(self)
        self.fl = True
        self.data_table = data_table

        self.container_widget_11 = self.findChild(QWidget, "buttonContainer_11")
        self.button_layout_11 = self.container_widget_11.layout()

        self.container_widget_12 = self.findChild(QWidget, "buttonContainer_12")
        self.button_layout_12 = self.container_widget_12.layout()

        self.container_widget_21 = self.findChild(QWidget, "buttonContainer_21")
        self.button_layout_21 = self.container_widget_21.layout()

        self.container_widget_22 = self.findChild(QWidget, "buttonContainer_22")
        self.button_layout_22 = self.container_widget_22.layout()

    def rang(self, data_table):
        self.clear_buttons()
        print('///////////')
        print(data_table)
        self.textEdit_2.clear()
        self.pushButton.clicked.connect(lambda: self.settxt(self.calculate_overlap, data_table))
        if len(data_table) == 2:
            self.add_button(self.button_layout_11, 'Стьюдент зав', 185, 30, had.studen_zav, data_table)
            self.add_button(self.button_layout_12, 'Стьюдент незав', 185, 30, had.studen_nez, data_table)
            self.add_button(self.button_layout_12, 'Манна-Уитни', 185, 30, had.mana_ui, data_table)
            self.add_button(self.button_layout_11, 'Вилкоксона', 185, 30, had.wilk, data_table)
            self.add_button(self.button_layout_22, 'Круаска Уолиса', 185, 30, had.kru, data_table)
        if len(data_table) > 2:
            self.add_button(self.button_layout_12, 'Круаска Уолиса', 185, 30, had.kru, data_table)
            self.add_button(self.button_layout_21, 'Стьюдент зав', 185, 30, had.studen_zav, data_table)
            self.add_button(self.button_layout_22, 'Стьюдент незав', 185, 30, had.studen_nez, data_table)
            self.add_button(self.button_layout_22, 'Манна-Уитни', 185, 30, had.mana_ui, data_table)
            self.add_button(self.button_layout_21, 'Вилкоксона', 185, 30, had.wilk, data_table)

    def settxt(self, func, data):
        self.textEdit_2.clear()
        self.textEdit_2.setStyleSheet("font-size: 16px;")
        self.textEdit_2.setPlainText(f'{func(data)}')

    def calculate_overlap(self, data_lists):
        """
        Функция для вычисления процента одинаковых данных между разными списками (дисперсиями).

        :param data_lists: Список списков данных, между которыми нужно найти пересечение.
        :return: Процент одинаковых данных в разных списках, как для каждой пары, так и общий.
        """
        if not data_lists or len(data_lists) < 2:
            raise ValueError("Необходимо передать как минимум два списка данных.")

        # Преобразуем каждый список в множество для упрощения поиска пересечения
        sets = [set(data) for data in data_lists]

        # Рассчитаем процент пересечения для каждой пары
        pair_overlap_percentages = []
        for i in range(len(sets)):
            for j in range(i + 1, len(sets)):
                intersection = sets[i].intersection(sets[j])
                total_elements = len(sets[i]) + len(sets[j])
                intersection_count = len(intersection)
                pair_overlap_percentage = (intersection_count / total_elements) * 100
                pair_overlap_percentages.append(f'pair={i + 1}/{j + 1}: {int(pair_overlap_percentage)}%')

        # Найдем пересечение всех множеств
        intersection_all = set.intersection(*sets)

        # Рассчитаем общий размер всех множеств
        total_elements_all = sum(len(s) for s in sets)

        # Рассчитаем общий процент пересечения
        intersection_count_all = len(intersection_all)
        overall_overlap_percentage = (intersection_count_all / total_elements_all) * 100

        # Формируем результат
        result = "\n".join(pair_overlap_percentages)
        result += f"\nОбщий процент пересечения: {int(overall_overlap_percentage)}%"

        return result

    def add_button(self, button_layout, label, width=None, height=None, action=None, data=None):
        print("Adding button:", label)
        button = QPushButton(label)

        # Установка размеров кнопки, если указаны
        if width and height:
            button.setFixedSize(width, height)

        # Подключаем действие к кнопке, если оно указано
        if action and data:
            print("Connecting button with action and data")
            button.clicked.connect(lambda: self.settxt(action, data))
        elif action:  # Если action передано, но data нет, можно добавить обработку
            print("Connecting button with action only")
            button.clicked.connect(lambda: self.settxt(action, None))

        # Добавляем кнопку в макет
        button_layout.addWidget(button)

    def closeEvent(self, event):
        """Отправляет сигнал при закрытии окна."""
        self.closed.emit()
        super().closeEvent(event)

    def clear_buttons(self):
        # Перебираем все элементы в макете
        while self.button_layout_11.count():
            item = self.button_layout_11.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()
        while self.button_layout_12.count():
            item = self.button_layout_12.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()
        while self.button_layout_21.count():
            item = self.button_layout_21.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()
        while self.button_layout_22.count():
            item = self.button_layout_22.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()

    def modify_data_table(self, data_table):
        text = f'Число дисперсий в ваших данных = {len(data_table)} \n\n  Вам подходят следующие критерии:'
        print(text)
        self.textEdit.setPlainText(text)  # Если это QTextEdit
        min_width = len(f'Число дисперсий в ваших данных = {len(data_table)}') * 10
        self.textEdit.setMinimumWidth(min_width)
        self.rang(data_table)


class SecondUIWindow(QWidget, SecondWindowForm):
    closed = pyqtSignal()

    def __init__(self, data_table):
        super(SecondUIWindow, self).__init__()
        self.setupUi(self)
        self.fl = True
        self.data_table = data_table

        self.container_widget_11 = self.findChild(QWidget, "buttonContainer_11")
        self.button_layout_11 = self.container_widget_11.layout()

        self.container_widget_12 = self.findChild(QWidget, "buttonContainer_12")
        self.button_layout_12 = self.container_widget_12.layout()

        self.container_widget_21 = self.findChild(QWidget, "buttonContainer_21")
        self.button_layout_21 = self.container_widget_21.layout()

        self.container_widget_22 = self.findChild(QWidget, "buttonContainer_22")
        self.button_layout_22 = self.container_widget_22.layout()

    def rang(self, data_table):
        self.clear_buttons()
        print('///////////')
        print(data_table)
        self.textEdit_3.clear()
        self.pushButton.clicked.connect(lambda: self.settxt(self.calculate_overlap, data_table))
        if len(data_table) == 2:
            self.add_button(self.button_layout_12, 'Фишера', 185, 30, had.fisher_test, data_table)
            self.add_button(self.button_layout_22, 'Бартлета', 185, 30, had.bart, data_table)
            self.add_button(self.button_layout_22, 'Левене', 185, 30, had.leve, data_table)
        if len(data_table) > 2:
            self.add_button(self.button_layout_12, 'Бартлета', 185, 30, had.bart, data_table)
            self.add_button(self.button_layout_12, 'Левене', 185, 30, had.leve, data_table)
            self.add_button(self.button_layout_22, 'Фишера', 185, 30, had.fisher_test, data_table)

    def settxt(self, func, data):
        self.textEdit_3.clear()
        self.textEdit_3.setStyleSheet("font-size: 16px;")
        self.textEdit_3.setPlainText(f'{func(data)}')

    def calculate_overlap(self, data_lists):
        """
        Функция для вычисления процента одинаковых данных между разными списками (дисперсиями).

        :param data_lists: Список списков данных, между которыми нужно найти пересечение.
        :return: Процент одинаковых данных в разных списках, как для каждой пары, так и общий.
        """
        if not data_lists or len(data_lists) < 2:
            raise ValueError("Необходимо передать как минимум два списка данных.")

        # Преобразуем каждый список в множество для упрощения поиска пересечения
        sets = [set(data) for data in data_lists]

        # Рассчитаем процент пересечения для каждой пары
        pair_overlap_percentages = []
        for i in range(len(sets)):
            for j in range(i + 1, len(sets)):
                intersection = sets[i].intersection(sets[j])
                total_elements = len(sets[i]) + len(sets[j])
                intersection_count = len(intersection)
                pair_overlap_percentage = (intersection_count / total_elements) * 100
                pair_overlap_percentages.append(f'pair={i + 1}/{j + 1}: {int(pair_overlap_percentage)}%')

        # Найдем пересечение всех множеств
        intersection_all = set.intersection(*sets)

        # Рассчитаем общий размер всех множеств
        total_elements_all = sum(len(s) for s in sets)

        # Рассчитаем общий процент пересечения
        intersection_count_all = len(intersection_all)
        overall_overlap_percentage = (intersection_count_all / total_elements_all) * 100

        # Формируем результат
        result = "\n".join(pair_overlap_percentages)
        result += f"\nОбщий процент пересечения: {int(overall_overlap_percentage)}%"

        return result

    def add_button(self, button_layout, label, width=None, height=None, action=None, data=None):
        print("Adding button:", label)
        button = QPushButton(label)

        # Установка размеров кнопки, если указаны
        if width and height:
            button.setFixedSize(width, height)

        # Подключаем действие к кнопке, если оно указано
        if action and data:
            print("Connecting button with action and data")
            button.clicked.connect(lambda: self.settxt(action, data))
        elif action:  # Если action передано, но data нет, можно добавить обработку
            print("Connecting button with action only")
            button.clicked.connect(lambda: self.settxt(action, None))

        # Добавляем кнопку в макет
        button_layout.addWidget(button)

    def closeEvent(self, event):
        """Отправляет сигнал при закрытии окна."""
        self.closed.emit()
        super().closeEvent(event)

    def clear_buttons(self):

        # Перебираем все элементы в макете
        while self.button_layout_11.count():
            item = self.button_layout_11.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()
        while self.button_layout_12.count():
            item = self.button_layout_12.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()
        while self.button_layout_21.count():
            item = self.button_layout_21.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()
        while self.button_layout_22.count():
            item = self.button_layout_22.takeAt(0)  # Удаляем элемент из макета
            widget = item.widget()  # Получаем виджет, если он существует
            if widget:
                widget.deleteLater()

    def modify_data_table(self, data_table):
        text = f'Число дисперсий в ваших данных = {len(data_table)} \n\n  Вам подходят следующие критерии:'
        print(text)
        self.textEdit.setPlainText(text)  # Если это QTextEdit
        min_width = len(f'Число дисперсий в ваших данных = {len(data_table)}') * 10
        self.textEdit.setMinimumWidth(min_width)
        self.rang(data_table)


class Heandler:
    def __init__(self):
        from scipy import stats
        self.stats = stats

    def studen_nez(self, data):
        """
        Выполняет t-тест для независимых выборок (двусторонний и односторонний) для всех возможных пар,
        если передано больше двух групп.
        """
        try:
            from itertools import combinations
            from scipy.stats import ttest_ind

            results = []
            results.append(f'Тест Стьюдента для независимых\n')
            # Если передано больше двух групп, вычисляем для всех пар
            if len(data) > 2:
                for (i, data1), (j, data2) in combinations(enumerate(data), 2):
                    # Двусторонний тест
                    stat, p_value_two_sided = ttest_ind(data1, data2, alternative='two-sided')

                    # Односторонние тесты
                    _, p_value_greater = ttest_ind(data1, data2, alternative='greater')
                    _, p_value_less = ttest_ind(data1, data2, alternative='less')

                    results.append(
                        f'Пара {i + 1} и {j + 1}:\n'
                        f'Статистика: {stat}\n'
                        f'p-значение (двустороннее): {p_value_two_sided}\n'
                        f'p-значение (одностороннее, больше): {p_value_greater}\n'
                        f'p-значение (одностороннее, меньше): {p_value_less}\n'
                    )
            else:
                # Обработка случая только для двух групп
                data1, data2 = data[0], data[1]

                # Двусторонний тест
                stat, p_value_two_sided = ttest_ind(data1, data2, alternative='two-sided')

                # Односторонние тесты
                _, p_value_greater = ttest_ind(data1, data2, alternative='greater')
                _, p_value_less = ttest_ind(data1, data2, alternative='less')

                results.append(
                    f'Статистика: {stat}\n'
                    f'p-значение (двустороннее): {p_value_two_sided}\n'
                    f'p-значение (одностороннее, больше): {p_value_greater}\n'
                    f'p-значение (одностороннее, меньше): {p_value_less}\n'
                )

            return "".join(results)

        except Exception as e:
            QMessageBox.critical(None, "Ошибка", f"Ошибка при выполнении t-теста для независимых выборок: {e}")
            return None


    def studen_zav(self, data):
        """
        Выполняет t-тест для зависимых выборок для всех возможных пар, если передано больше двух групп.
        """
        try:
            from itertools import combinations
            from scipy.stats import ttest_rel

            results = []
            results.append(f'Тест Стьюдента для зависимых\n')
            if len(data) > 2:
                for (i, data1), (j, data2) in combinations(enumerate(data), 2):
                    stat, p_value = ttest_rel(data1, data2)
                    results.append(
                        f'Пара {i + 1} и {j + 1}:\n'
                        f'Статистика: {stat}\n'
                        f'p-значение: {p_value}\n'
                    )
            else:
                data1, data2 = data[0], data[1]
                stat, p_value = ttest_rel(data1, data2)
                results.append(
                    f'Статистика: {stat}\n'
                    f'p-значение: {p_value}\n'
                )

            return "".join(results)

        except Exception as e:
            return f"Ошибка при выполнении t-теста для зависимых выборок: {e}"

    def mana_ui(self, data):
        """
        Выполняет тест Манна-Уитни для всех возможных пар, если передано больше двух групп.
        """
        try:
            from itertools import combinations
            from scipy.stats import mannwhitneyu

            results = []
            results.append(f'Тест Манна-Уитни\n')
            if len(data) > 2:
                for (i, data1), (j, data2) in combinations(enumerate(data), 2):
                    stat, p_value = mannwhitneyu(data1, data2)
                    results.append(
                        f'Пара {i + 1} и {j + 1}:\n'
                        f'Статистика: {stat}\n'
                        f'p-значение: {p_value}\n'
                    )
            else:
                data1, data2 = data[0], data[1]
                stat, p_value = mannwhitneyu(data1, data2)
                results.append(
                    f'Статистика: {stat}\n'
                    f'p-значение: {p_value}\n'
                )

            return "".join(results)

        except Exception as e:
            return f"Ошибка при выполнении теста Манна-Уитни: {e}"

    def wilk(self, data):
        """
        Выполняет тест Уилкоксона для всех возможных пар, если передано больше двух групп.
        """
        try:
            from itertools import combinations
            from scipy.stats import wilcoxon

            results = []
            results.append(f'Тест Уилкоксона\n')
            if len(data) > 2:
                for (i, data1), (j, data2) in combinations(enumerate(data), 2):
                    stat, p_value = wilcoxon(data1, data2)
                    results.append(
                        f'Пара {i + 1} и {j + 1}:\n'
                        f'Статистика: {stat}\n'
                        f'p-значение: {p_value}\n'
                    )
            else:
                data1, data2 = data[0], data[1]
                stat, p_value = wilcoxon(data1, data2)
                results.append(
                    f'Статистика: {stat}\n'
                    f'p-значение: {p_value}\n'
                )

            return "".join(results)

        except Exception as e:
            return f"Ошибка при выполнении теста Уилкоксона: {e}"

    def kru(self, data):
        """
        Выполняет тест Крускала-Уоллиса.
        """
        try:
            stat, p_value = self.stats.kruskal(*data)
            return (
                f'Тест Крускала-Уоллиса\n'
                f'Статистика: {stat}\n'
                f'p-значение: {p_value}\n'
            )
        except Exception as e:
            return f"Ошибка при выполнении теста Крускала-Уоллиса: {e}"

    def fisher_test(self, data):
        """
        Выполняет F-тест для всех возможных пар, если передано больше двух выборок.
        """
        try:
            import numpy as np
            from scipy.stats import f
            from itertools import combinations

            results = []
            results.append(f'Тест Фишера\n')
            if len(data) > 2:
                for (i, data1), (j, data2) in combinations(enumerate(data), 2):
                    var1 = np.var(data1, ddof=1)
                    var2 = np.var(data2, ddof=1)
                    F_statistic = var1 / var2 if var1 > var2 else var2 / var1
                    p_value = 1 - f.cdf(F_statistic, len(data1) - 1, len(data2) - 1)

                    results.append(
                        f'Пара {i + 1} и {j + 1}:\n'
                        f'F-статистика: {F_statistic}\n'
                        f'p-значение: {p_value}\n'
                    )
            else:
                data1, data2 = data[0], data[1]
                var1 = np.var(data1, ddof=1)
                var2 = np.var(data2, ddof=1)
                F_statistic = var1 / var2 if var1 > var2 else var2 / var1
                p_value = 1 - f.cdf(F_statistic, len(data1) - 1, len(data2) - 1)

                results.append(
                    f'F-статистика: {F_statistic}\n'
                    f'p-значение: {p_value}\n'
                )

            return "".join(results)

        except Exception as e:
            return f"Ошибка при выполнении F-теста: {e}"

    def bart(self, data):
        """
        Выполняет тест Бартлетта.
        """
        try:
            stat, p_value = self.stats.bartlett(*data)
            return (
                f'Тест Бартлетта\n'
                f'Статистика: {stat}\n'
                f'p-значение: {p_value}\n'
            )
        except Exception as e:
            return f"Ошибка при выполнении теста Бартлетта: {e}"

    def leve(self, data):
        """
        Выполняет тест Левена.
        """
        try:
            stat, p_value = self.stats.levene(*data)
            return (
                f'Тест Левена\n'
                f'Статистика: {stat}\n'
                f'p-значение: {p_value}\n'
            )
        except Exception as e:
            return f"Ошибка при выполнении теста Левена: {e}"




if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    had = Heandler()
    w = Ui(had)
    w.show()
    sys.exit(app.exec())
